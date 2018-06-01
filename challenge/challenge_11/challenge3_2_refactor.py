# -*-coding: utf-8 -*-
from openpyxl import load_workbook, Workbook
import datetime

def combine():
    wb = load_workbook('courses.xlsx')
    students = wb['students']
    time = wb.get_sheet_by_name('time')
    combine = wb.create_sheet('combine')
    combine.append(['创建时间', '课程名称', '学习人数', '学习时间'])
    for a, b in zip(list(students.values)[1:], list(time.values)[1:]):
        combine.append(list(a)+[b[-1],])
    wb.save('courses.xlsx')

def split():
    wb = load_workbook('courses.xlsx')
    combine = wb['combine']
    year_set = set()
    content_list = list(combine.values)[1:]
    for row in content_list:
        year_set.add(row[0].strftime('%Y'))
    for year in year_set:
        new_wb = Workbook()
        ws = new_wb.active
        ws.title = year
        for row in content_list:
            if row[0].strftime('%Y') == year:
                ws.append(row)
        new_wb.save(str(year)+'.xlsx')

if __name__ == '__main__':
    combine()
    split()
