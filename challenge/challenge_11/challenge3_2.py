# -*-coding: utf-8 -*-
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
import datetime

def combine():
    wb = load_workbook(filename='courses.xlsx')
    wb.active
    students = wb.get_sheet_by_name('students')
    time = wb.get_sheet_by_name('time')
    combine = wb.create_sheet('combine')
    list_students = []
    list_time = []
    list_combine = []
    for row in students.rows:
        row_students = []
        for row_cell in row:
            row_students.append(row_cell.value)
        list_students.append(row_students)
    for row in time.rows:
        row_time = []
        for row_cell in row[::2]:
            row_time.append(row_cell.value)
        list_time.append(row_time)
    for l_student in list_students[1:]:
        for l_time in list_time[1:]:
            if l_time[0] == l_student[0]:
                l_student[0] = l_student[0].strftime("%Y-%m-%d")
                l_student.append(l_time[1])
                list_combine.append(l_student)
    list_combine.sort()
    list_students[0].append(list_time[0][-1])
    list_combine.insert(0,list_students[0])
    i = 1
    for line in list_combine:
        for col in range(1, len(line) + 1):
            col_num = get_column_letter(col)
            combine.cell('%s%s'%(col_num, i)).value = line[col-1]
        i = i + 1
    wb.save('courses.xlsx')

def split():
    wb = load_workbook(filename='courses.xlsx')
    wb.active
    combine = wb.get_sheet_by_name('combine')
    list_combine = []
    for row in combine.rows:
        row_combine = []
        for row_cell in row:
            row_combine.append(row_cell.value)
        list_combine.append(row_combine)
    year_content = {}
    for combine in list_combine[1:]:
        year = datetime.datetime.strptime(combine[0], "%Y-%m-%d").year
        if year in year_content.keys():
            year_content[year].append(combine)
        else:
            year_content[year] = [combine,]
    for key, value in year_content.items():
        print(key, value)
        new_wb = Workbook()
        year_sheet = new_wb.create_sheet(str(year))
        i = 1
        for line in value:
            for col in range(1, len(line) + 1):
                col_num = get_column_letter(col)
                year_sheet.cell('%s%s'%(col_num, i)).value = line[col-1]
            i = i + 1
        new_wb.save(str(year)+'.xlsx')



if __name__ == '__main__':
#    combine()
    split()
